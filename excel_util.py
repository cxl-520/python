import openpyxl


def read_excel(file_path, sheet_name="Sheet1"):
    wb = openpyxl.load_workbook(file_path)
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        sheet = wb.active

    rows = sheet.max_row
    cols = sheet.max_column
    data = []

    headers = [sheet.cell(1, col).value for col in range(1, cols + 1)]
    if not any(headers):
        return data

    for row in range(2, rows + 1):
        row_data = {}
        empty = True
        for col in range(1, cols + 1):
            key = headers[col - 1]
            value = sheet.cell(row, col).value
            if value is not None:
                empty = False
            row_data[key] = value
        if empty:
            continue
        data.append(row_data)
    return data


if __name__ == "__main__":
    test_path = r"C:\Users\Administrator\PycharmProjects\PythonProject\data\data.xlsx"
    print(read_excel(test_path))
