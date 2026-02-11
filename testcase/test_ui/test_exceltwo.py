import os
import openpyxl
import pytest
import selenium
from ddt import ddt, data  # 仅用data，不用unpack


# ========== 第一步：读取Excel + 打印日志（确认数据读取成功） ==========
def read_excel(file_path, sheet_name="Sheet1"):
    print("===== 第一步：开始读取Excel =====")
    # 1. 校验文件存在
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel文件不存在：{file_path}")
    print(f"✅ Excel文件路径正确：{file_path}")

    # 2. 读取Excel
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    test_data = []
    # 按列读取（第1列=username，第2列=password）
    for row in range(2, sheet.max_row + 1):
        row_data = {
            "username": sheet.cell(row, 1).value,
            "password": sheet.cell(row, 2).value
        }
        test_data.append(row_data)
        print(f"📝 读取第{row}行数据：{row_data}")

    # 3. 打印最终读取结果
    print(f"✅ 读取完成，总数据条数：{len(test_data)}")
    print(f"📊 最终测试数据：{test_data}")
    return test_data


# ========== 第二步：执行读取 + 强制校验数据 ==========
excel_path = r"/data/data.xlsx"
test_data = read_excel(excel_path, sheet_name="Sheet1")

# 强制校验：数据必须是非空列表，且每个元素是字典
assert isinstance(test_data, list), f"❌ 数据格式错误，不是列表：{type(test_data)}"
assert len(test_data) > 0, "❌ 测试数据为空！"
for idx, item in enumerate(test_data):
    assert isinstance(item, dict), f"❌ 第{idx}条数据不是字典：{item}"
    assert "username" in item and "password" in item, f"❌ 第{idx}条数据缺少key：{item}"
print("===== 第二步：数据校验通过 =====")


# ========== 第三步：DDT传参 + 打印日志（确认数据传入测试方法） ==========
@ddt
class TestExcel():
    print("===== 第三步：DDT装饰类完成，准备传参 =====")

    @data(*test_data)  # 解包列表，逐个传入字典
    def test_read_excel(self, data_item):
        # 打印传入的参数（核心：确认数据是否传进来）
        print("\n===== 测试方法内：接收参数 =====")
        print(f"✅ 传入的原始数据：{data_item}")

        # 读取参数（加容错）
        username = data_item.get("username", "❌ 未获取到")
        password = data_item.get("password", "❌ 未获取到")
        print(f"✅ 解析后：账号={username}，密码={password}")

        # 断言兜底
        assert username != "❌ 未获取到", "用户名未传进来！"
        assert password != "❌ 未获取到", "密码未传进来！"


if __name__ == "__main__":
    # 关键：-s显示所有print日志，-x出错立即停止，方便定位
    print("===== 开始运行测试脚本 =====")
    pytest.main(["-v", "-s", "-x", __file__])
    import os

    import openpyxl
    import unittest
    import selenium
    from ddt import ddt, data
    from excel_util import read_excel

    test_data = read_excel(r"/data/data.xlsx", sheet_name="Sheet1")


    @ddt
    class TestExcel(unittest.TestCase):
        @data(*test_data)
        def test_read_excel(self, data_json):
            clean_keys = data_json
            username = clean_keys.get("username")
            password = clean_keys.get("password")
            print(f"账号是{username},密码是{password}")


    if __name__ == "__main__":
        unittest.main(verbosity=2)
